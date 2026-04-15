from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

import scrape_tripadvisor_attractions as scraper


class ScraperUtilitiesTestCase(unittest.TestCase):
    def test_load_previous_outputs_returns_none_for_missing_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            record = Path(tmp_dir) / "missing.json"
            excel, video = scraper.load_previous_outputs(record)
            self.assertIsNone(excel)
            self.assertIsNone(video)

    def test_load_previous_outputs_parses_json_payload(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            record = Path(tmp_dir) / "latest.json"
            payload = {
                "excel": str(Path(tmp_dir) / "latest.xlsx"),
                "video": str(Path(tmp_dir) / "latest.webm"),
            }
            record.write_text(json.dumps(payload), encoding="utf-8")

            excel, video = scraper.load_previous_outputs(record)

            self.assertEqual(excel, Path(payload["excel"]))
            self.assertEqual(video, Path(payload["video"]))

    def test_load_previous_outputs_returns_none_for_invalid_json(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            record = Path(tmp_dir) / "latest.json"
            record.write_text("{ invalid json }", encoding="utf-8")

            excel, video = scraper.load_previous_outputs(record)

            self.assertIsNone(excel)
            self.assertIsNone(video)

    def test_remove_previous_outputs_only_deletes_stale_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            previous_excel = tmp_path / "old.xlsx"
            previous_video = tmp_path / "old.webm"
            current_excel = tmp_path / "current.xlsx"
            current_video = tmp_path / "current.webm"

            for path in (previous_excel, previous_video, current_excel, current_video):
                path.write_text("placeholder", encoding="utf-8")

            scraper.remove_previous_outputs(
                previous_excel=previous_excel,
                previous_video=previous_video,
                current_excel=current_excel,
                current_video=current_video,
            )

            self.assertFalse(previous_excel.exists())
            self.assertFalse(previous_video.exists())
            self.assertTrue(current_excel.exists())
            self.assertTrue(current_video.exists())

    def test_persist_latest_outputs_writes_absolute_paths(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            record = tmp_path / "latest.json"
            excel = tmp_path / "result.xlsx"
            video = tmp_path / "result.webm"

            scraper.persist_latest_outputs(record, excel, video)
            persisted = json.loads(record.read_text(encoding="utf-8"))

            self.assertEqual(persisted["excel"], str(excel.resolve()))
            self.assertEqual(persisted["video"], str(video.resolve()))

    @mock.patch("scrape_tripadvisor_attractions.shutil.which", return_value=None)
    def test_trim_video_from_offset_falls_back_when_ffmpeg_missing(
        self, _mock_which: mock.Mock
    ) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            source = tmp_path / "raw.webm"
            target = tmp_path / "trimmed.webm"
            source.write_text("raw", encoding="utf-8")

            success = scraper.trim_video_from_offset(source, target, start_offset_seconds=1.2)

            self.assertFalse(success)
            self.assertFalse(target.exists())


class ScraperCliAndOutputTestCase(unittest.TestCase):
    def test_parse_arguments_accepts_custom_values(self) -> None:
        with mock.patch.object(
            sys,
            "argv",
            [
                "scrape_tripadvisor_attractions.py",
                "--start-page",
                "6",
                "--pages",
                "3",
                "--output",
                "custom.xlsx",
                "--video",
                "custom.webm",
                "--headed",
            ],
        ):
            args = scraper.parse_arguments()

        self.assertEqual(args.start_page, 6)
        self.assertEqual(args.pages, 3)
        self.assertEqual(args.output, "custom.xlsx")
        self.assertEqual(args.video, "custom.webm")
        self.assertTrue(args.headed)

    def test_save_to_excel_keeps_expected_sheet_and_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output = Path(tmp_dir) / "attractions.xlsx"
            items = [
                scraper.AttractionItem(rank=1, name="北京故宫"),
                scraper.AttractionItem(rank=2, name="长城"),
            ]

            scraper.save_to_excel(items, output)

            workbook = load_workbook(output)
            self.assertEqual(workbook.sheetnames, ["景点"])
            sheet = workbook["景点"]
            self.assertEqual(sheet.cell(row=1, column=1).value, "编号")
            self.assertEqual(sheet.cell(row=1, column=2).value, "景点名称")
            self.assertEqual(sheet.cell(row=2, column=1).value, 1)
            self.assertEqual(sheet.cell(row=2, column=2).value, "北京故宫")
            self.assertEqual(sheet.cell(row=3, column=1).value, 2)
            self.assertEqual(sheet.cell(row=3, column=2).value, "长城")
            workbook.close()

    def test_run_scraper_rejects_invalid_pages(self) -> None:
        with self.assertRaisesRegex(ValueError, "pages 必须大于 0"):
            scraper.run_scraper(
                url=scraper.DEFAULT_URL,
                pages=0,
                start_page=1,
                output=Path("output.xlsx"),
                video_output=Path("output.webm"),
                headed=False,
            )

    def test_run_scraper_rejects_invalid_start_page(self) -> None:
        with self.assertRaisesRegex(ValueError, "start_page 必须大于 0"):
            scraper.run_scraper(
                url=scraper.DEFAULT_URL,
                pages=1,
                start_page=0,
                output=Path("output.xlsx"),
                video_output=Path("output.webm"),
                headed=False,
            )


if __name__ == "__main__":
    unittest.main()
